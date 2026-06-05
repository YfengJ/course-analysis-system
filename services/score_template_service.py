from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class ScoreTemplateService:
    """按当前课程配置生成可直接上传的简化成绩模板。"""

    SHEET_NAME = "达成度计算表"
    TEMPLATE_ROW_COUNT = 506

    HEADER_FILL = PatternFill("solid", fgColor="D9E6F8")
    SUB_HEADER_FILL = PatternFill("solid", fgColor="EEF3FB")
    INPUT_FILL = PatternFill("solid", fgColor="FFFFFF")
    FORMULA_FILL = PatternFill("solid", fgColor="F1F3F5")
    NOTE_FILL = PatternFill("solid", fgColor="FFF4D6")
    THIN_BORDER = Border(
        left=Side(style="thin", color="9AA7B4"),
        right=Side(style="thin", color="9AA7B4"),
        top=Side(style="thin", color="9AA7B4"),
        bottom=Side(style="thin", color="9AA7B4"),
    )

    @classmethod
    def build_course_template(cls, course, output_path: Path) -> Path:
        objectives = sorted(course.objectives, key=lambda item: item.sequence)
        if not objectives:
            raise ValueError("当前课程还没有课程目标，请先导入教学大纲或维护课程目标后再下载成绩模板。")

        objective_blocks = []
        for objective in objectives:
            weights = sorted(objective.assessment_weights, key=lambda item: item.assessment.sequence)
            if not weights:
                continue
            objective_blocks.append((objective, weights))

        if not objective_blocks:
            raise ValueError("当前课程还没有课程目标与考核项的权重配置，请先导入教学大纲后再下载成绩模板。")

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = cls.SHEET_NAME

        total_column = cls._write_headers(sheet, course, objective_blocks)
        cls._write_formula_rows(sheet, objective_blocks, total_column)
        cls._style_sheet(sheet, total_column)

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        return output_path

    @classmethod
    def _write_headers(cls, sheet, course, objective_blocks):
        sheet["A1"] = f"{course.name}课程目标达成度成绩导入模板"
        sheet["A1"].font = Font(name="Microsoft YaHei", size=16, bold=True)
        sheet["A2"] = "填写说明：只需从第7行开始填写学号、姓名、班级和各课程目标下的分项得分；灰色达成度列为辅助公式，可不手工填写。"
        sheet["A2"].fill = cls.NOTE_FILL
        sheet["A3"] = "学生基本信息"
        sheet["A4"] = "序号"
        sheet["B4"] = "学号"
        sheet["C4"] = "姓名"
        sheet["D4"] = "班级"
        sheet["A5"] = "说明"
        sheet["B5"] = "必填"
        sheet["C5"] = "建议填写"
        sheet["D5"] = "建议填写"
        sheet["A6"] = "从下方开始填写"

        column = 5
        attainment_columns = []
        for objective, weights in objective_blocks:
            start_column = column
            sheet.cell(row=3, column=start_column, value=objective.title)
            for weight in weights:
                sheet.cell(row=4, column=column, value=weight.assessment.name)
                sheet.cell(row=5, column=column, value=float(weight.weight_score or 0))
                sheet.cell(row=6, column=column, value="分项满分")
                column += 1

            attainment_column = column
            sheet.cell(row=4, column=attainment_column, value="达成度(%)")
            sheet.cell(row=5, column=attainment_column, value=float(objective.weight or 0))
            sheet.cell(row=6, column=attainment_column, value="系统自动计算")
            attainment_columns.append((attainment_column, objective.weight or 0))

            if attainment_column > start_column:
                sheet.merge_cells(start_row=3, start_column=start_column, end_row=3, end_column=attainment_column)
            column += 1

        total_column = column
        sheet.cell(row=3, column=total_column, value="课程总目标")
        sheet.cell(row=4, column=total_column, value="总达成度(%)")
        sheet.cell(row=5, column=total_column, value=100)
        sheet.cell(row=6, column=total_column, value="系统自动计算")

        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_column)
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_column)
        sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=4)
        return total_column

    @classmethod
    def _write_formula_rows(cls, sheet, objective_blocks, total_column):
        objective_start_column = 5
        attainment_columns = []
        column = objective_start_column
        for objective, weights in objective_blocks:
            first_score_column = column
            last_score_column = column + len(weights) - 1
            attainment_column = last_score_column + 1
            attainment_columns.append(attainment_column)
            column = attainment_column + 1

            first_letter = get_column_letter(first_score_column)
            last_letter = get_column_letter(last_score_column)
            attainment_letter = get_column_letter(attainment_column)
            for row in range(7, cls.TEMPLATE_ROW_COUNT + 1):
                sheet.cell(row=row, column=1, value=f'=IF($B{row}="","",ROW()-6)')
                sheet.cell(
                    row=row,
                    column=attainment_column,
                    value=f'=IF($B{row}="","",ROUND(SUM({first_letter}{row}:{last_letter}{row})/{attainment_letter}$5*100,2))',
                )

        for row in range(7, cls.TEMPLATE_ROW_COUNT + 1):
            terms = [
                f"{get_column_letter(attainment_column)}{row}*{get_column_letter(attainment_column)}$5/100"
                for attainment_column in attainment_columns
            ]
            sheet.cell(
                row=row,
                column=total_column,
                value=f'=IF($B{row}="","",ROUND({" + ".join(terms)},2))',
            )

    @classmethod
    def _style_sheet(cls, sheet, total_column):
        sheet.freeze_panes = "E7"
        sheet.sheet_view.showGridLines = False

        widths = {
            1: 10,
            2: 16,
            3: 14,
            4: 16,
        }
        for column in range(1, total_column + 1):
            sheet.column_dimensions[get_column_letter(column)].width = widths.get(column, 13)

        for row in range(1, cls.TEMPLATE_ROW_COUNT + 1):
            sheet.row_dimensions[row].height = 24

        for row in range(1, 7):
            for column in range(1, total_column + 1):
                cell = sheet.cell(row=row, column=column)
                cell.font = Font(name="Microsoft YaHei", size=10, bold=row in {1, 3, 4})
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if row in {3, 4}:
                    cell.fill = cls.HEADER_FILL
                elif row in {5, 6}:
                    cell.fill = cls.SUB_HEADER_FILL
                cell.border = cls.THIN_BORDER

        sheet["A1"].font = Font(name="Microsoft YaHei", size=16, bold=True)
        sheet["A2"].font = Font(name="Microsoft YaHei", size=10, color="7A4E00")

        formula_headers = {"达成度(%)", "总达成度(%)"}
        for row in range(7, cls.TEMPLATE_ROW_COUNT + 1):
            for column in range(1, total_column + 1):
                cell = sheet.cell(row=row, column=column)
                header = sheet.cell(row=4, column=column).value
                cell.font = Font(name="Microsoft YaHei", size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.fill = cls.FORMULA_FILL if header in formula_headers or column == 1 else cls.INPUT_FILL
                cell.border = cls.THIN_BORDER

        for column in range(5, total_column + 1):
            sheet.column_dimensions[get_column_letter(column)].width = 14
